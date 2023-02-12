from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('SPEED.xlsx')
ws = wb.active
Resultado_suma = 0

#Odd cells get multyply by 4 and even cells get multiply by 2.
for row in range(3, 40981):
    if row % 2 == 1:
        for col in [2]:
            char = get_column_letter(col)
            Resultado_suma += (ws[char + str(row)].value) * 4
    else:
        for col in [2]:
            char = get_column_letter(col)
            Resultado_suma += (ws[char + str(row)].value) * 2

#Sum first and last cells
Resultado_suma += ws["B2"].value + ws["B40981"].value

#Get integration limits and get delta x
Delta_x = (ws["A40981"].value - ws["A2"].value) / 40980

Resultado_final = (Resultado_suma * Delta_x) / 3
print(Resultado_final)
 
"""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⣿⣶⣄⣀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⢀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣦⣄⣀⡀⣠⣾⡇⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀
⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⢿⣿⣿⡇⠀⠀⠀⠀
⠀⣶⣿⣦⣜⣿⣿⣿⡟⠻⣿⣿⣿⣿⣿⣿⣿⡿⢿⡏⣴⣺⣦⣙⣿⣷⣄⠀⠀⠀
⠀⣯⡇⣻⣿⣿⣿⣿⣷⣾⣿⣬⣥⣭⣽⣿⣿⣧⣼⡇⣯⣇⣹⣿⣿⣿⣿⣧⠀⠀
⠀⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠸⣿⣿⣿⣿⣿⣿⣿⣷⠀
""" 
